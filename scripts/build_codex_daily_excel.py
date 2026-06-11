from __future__ import annotations

import json
import math
import ssl
import shutil
import subprocess
import time
import urllib.request
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parents[1]
OUTPUT_DIR = BASE_DIR / "output"
BEIJING_NOW = datetime.now(ZoneInfo("Asia/Shanghai"))
REPORT_DATE = BEIJING_NOW.strftime("%Y-%m-%d")
END_DATE = BEIJING_NOW.strftime("%Y%m%d")
START_DATE = "20240401"
OUTPUT_FILE = OUTPUT_DIR / f"codex_daily_observation_{REPORT_DATE}.xlsx"
DATA_AUDIT: dict[str, dict[str, Any]] = {}
EASTMONEY_AVAILABLE: bool | None = None


@dataclass
class Item:
    kind: str
    code: str
    name: str
    secid: str
    sector: str = ""
    buy_price: float | None = None


ITEMS = [
    Item("stock", "605111", "新洁能", "1.605111", "半导体", 66.5),
    Item("stock", "600703", "三安光电", "1.600703", "光学光电子", 16.0),
    Item("stock", "002245", "蔚蓝锂芯", "0.002245", "电池", 21.0),
    Item("stock", "600522", "中天科技", "1.600522", "通信设备", 46.0),
    Item("index", "000300", "沪深300", "1.000300"),
    Item("index", "000852", "中证1000", "1.000852"),
    Item("index", "399006", "创业板指", "0.399006"),
    Item("sector", "BK1036", "半导体", "90.BK1036"),
    Item("sector", "BK1032", "光学光电子", "90.BK1032"),
    Item("sector", "BK1033", "电池", "90.BK1033"),
    Item("sector", "BK0736", "通信设备", "90.BK0736"),
]

SECTOR_BASKETS = {
    "半导体": ["688981", "603501", "600584", "002371", "688041"],
    "光学光电子": ["000725", "002456", "300433", "002273", "600707"],
    "电池": ["300750", "002074", "300014", "002460", "002709"],
    "通信设备": ["000063", "300308", "300502", "600498", "002281"],
}


def is_number(value: Any) -> bool:
    try:
        return value is not None and math.isfinite(float(value))
    except (TypeError, ValueError):
        return False


def avg(values: list[float]) -> float | None:
    clean = [float(v) for v in values if is_number(v)]
    if not clean:
        return None
    return sum(clean) / len(clean)


def tail_avg(values: list[float], n: int) -> float | None:
    if len(values) < n:
        return None
    return avg(values[-n:])


def pct(value: float | None) -> float | None:
    if not is_number(value):
        return None
    return float(value)


def record_data_audit(item: Item, source: str, rows: list[dict[str, Any]], note: str = "") -> None:
    DATA_AUDIT[item.code] = {
        "类别": item.kind,
        "代码": item.code,
        "名称": item.name,
        "数据来源": source,
        "是否缓存": "是" if source == "Excel缓存" else "否",
        "最新行情日期": rows[-1]["date"] if rows else None,
        "行情行数": len(rows),
        "说明": note,
    }


def tencent_symbol(item: Item) -> str:
    prefix = "sh" if item.secid.startswith("1.") else "sz"
    return f"{prefix}{item.code}"


def stock_item(code: str) -> Item:
    market = "1" if code.startswith(("5", "6", "9")) else "0"
    return Item("stock", code, code, f"{market}.{code}")


def fetch_tencent_klines(item: Item) -> list[dict[str, Any]]:
    if item.kind not in {"stock", "index"}:
        return []

    symbol = tencent_symbol(item)
    url = f"https://web.ifzq.gtimg.cn/appstock/app/fqkline/get?param={symbol},day,,,1023,qfq"
    payload = fetch_json(url)
    node = (payload.get("data") or {}).get(symbol) or {}
    lines = node.get("qfqday") or node.get("day") or []
    rows: list[dict[str, Any]] = []
    previous_close: float | None = None
    for line in lines:
        if len(line) < 6 or str(line[0]) < START_DATE[:4] + "-" + START_DATE[4:6] + "-" + START_DATE[6:]:
            continue
        open_price = float(line[1])
        close = float(line[2])
        high = float(line[3])
        low = float(line[4])
        volume = float(line[5])
        change_amount = close - previous_close if previous_close else 0.0
        pct_change = change_amount / previous_close if previous_close else 0.0
        amplitude = (high - low) / previous_close if previous_close else 0.0
        rows.append(
            {
                "date": str(line[0]),
                "open": open_price,
                "close": close,
                "high": high,
                "low": low,
                "volume": volume,
                "amount": 0.0,
                "amplitude": amplitude,
                "pct_change": pct_change,
                "change_amount": change_amount,
                "turnover": None,
            }
        )
        previous_close = close
    return rows


def fetch_sector_basket_klines(item: Item) -> list[dict[str, Any]]:
    constituents = SECTOR_BASKETS.get(item.name) or []
    if not constituents:
        return []

    daily_returns: dict[str, list[float]] = {}
    daily_volumes: dict[str, float] = {}
    for code in constituents:
        rows = fetch_tencent_klines(stock_item(code))
        if not rows:
            continue
        for row in rows:
            daily_returns.setdefault(row["date"], []).append(float(row["pct_change"]))
            daily_volumes[row["date"]] = daily_volumes.get(row["date"], 0.0) + float(row["volume"])

    rows: list[dict[str, Any]] = []
    synthetic_close = 1000.0
    for date_value in sorted(daily_returns):
        returns = daily_returns[date_value]
        if len(returns) < max(3, len(constituents) - 1):
            continue
        daily_return = sum(returns) / len(returns)
        previous_close = synthetic_close
        synthetic_close = previous_close * (1 + daily_return)
        open_price = previous_close
        rows.append(
            {
                "date": date_value,
                "open": open_price,
                "close": synthetic_close,
                "high": max(open_price, synthetic_close),
                "low": min(open_price, synthetic_close),
                "volume": daily_volumes.get(date_value, 0.0),
                "amount": 0.0,
                "amplitude": abs(daily_return),
                "pct_change": daily_return,
                "change_amount": synthetic_close - previous_close,
                "turnover": None,
            }
        )
    return rows


def fetch_klines(item: Item) -> list[dict[str, Any]]:
    global EASTMONEY_AVAILABLE

    url = (
        "https://push2his.eastmoney.com/api/qt/stock/kline/get"
        f"?secid={item.secid}"
        "&fields1=f1,f2,f3,f4,f5,f6"
        "&fields2=f51,f52,f53,f54,f55,f56,f57,f58,f59,f60,f61"
        "&klt=101&fqt=1"
        f"&beg={START_DATE}&end={END_DATE}"
    )
    eastmoney_error: Exception | None = None
    if EASTMONEY_AVAILABLE is not False:
        try:
            payload = fetch_json(url)
            klines = (payload.get("data") or {}).get("klines") or []
            if klines:
                rows: list[dict[str, Any]] = []
                for line in klines:
                    parts = line.split(",")
                    rows.append(
                        {
                            "date": parts[0],
                            "open": float(parts[1]),
                            "close": float(parts[2]),
                            "high": float(parts[3]),
                            "low": float(parts[4]),
                            "volume": float(parts[5]),
                            "amount": float(parts[6]),
                            "amplitude": float(parts[7]) / 100,
                            "pct_change": float(parts[8]) / 100,
                            "change_amount": float(parts[9]),
                            "turnover": float(parts[10]) / 100 if parts[10] != "-" else None,
                        }
                    )
                EASTMONEY_AVAILABLE = True
                record_data_audit(item, "东方财富前复权", rows)
                return rows
            eastmoney_error = RuntimeError("东方财富返回空行情")
            EASTMONEY_AVAILABLE = False
        except Exception as exc:  # noqa: BLE001
            eastmoney_error = exc
            EASTMONEY_AVAILABLE = False
    else:
        eastmoney_error = RuntimeError("东方财富已在本次运行中失败，跳过重复请求")

    if item.kind in {"stock", "index"}:
        try:
            rows = fetch_tencent_klines(item)
            if rows:
                record_data_audit(item, "腾讯前复权", rows, f"东方财富失败：{eastmoney_error}")
                return rows
        except Exception as exc:  # noqa: BLE001
            eastmoney_error = RuntimeError(f"东方财富失败：{eastmoney_error}；腾讯失败：{exc}")

    if item.kind == "sector":
        try:
            rows = fetch_sector_basket_klines(item)
            if rows:
                note = f"东方财富失败，使用腾讯成分股等权代理：{','.join(SECTOR_BASKETS[item.name])}"
                record_data_audit(item, "腾讯成分股等权代理", rows, note)
                return rows
        except Exception as exc:  # noqa: BLE001
            eastmoney_error = RuntimeError(f"东方财富失败：{eastmoney_error}；板块代理失败：{exc}")

    cached_rows = load_cached_klines(item)
    if cached_rows:
        record_data_audit(item, "Excel缓存", cached_rows, f"联网行情失败：{eastmoney_error}")
        return cached_rows
    raise RuntimeError(f"No kline data for {item.name} {item.secid}: {eastmoney_error}")


def load_cached_klines(item: Item) -> list[dict[str, Any]]:
    candidates = sorted(OUTPUT_DIR.glob("codex_daily_observation_*.xlsx"), reverse=True)
    if OUTPUT_FILE.exists():
        candidates.insert(0, OUTPUT_FILE)
    if (BASE_DIR / "latest_report.xlsx").exists():
        candidates.append(BASE_DIR / "latest_report.xlsx")
    candidates.extend(sorted(BASE_DIR.glob("codex_daily_observation_*.xlsx"), reverse=True))

    seen: set[Path] = set()
    for candidate in candidates:
        if candidate in seen or not candidate.exists():
            continue
        seen.add(candidate)
        try:
            wb = load_workbook(candidate, data_only=True, read_only=True)
            if "原始日线" not in wb.sheetnames:
                continue
            ws = wb["原始日线"]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            rows: list[dict[str, Any]] = []
            for values in ws.iter_rows(min_row=2, values_only=True):
                record = dict(zip(headers, values))
                if str(record.get("代码", "")).zfill(6) != item.code:
                    continue
                rows.append(
                    {
                        "date": record["date"],
                        "open": float(record["open"]),
                        "close": float(record["close"]),
                        "high": float(record["high"]),
                        "low": float(record["low"]),
                        "volume": float(record["volume"]),
                        "amount": float(record.get("amount") or 0),
                        "amplitude": float(record.get("amplitude") or 0),
                        "pct_change": float(record.get("pct_change") or 0),
                        "change_amount": float(record.get("change_amount") or 0),
                        "turnover": float(record["turnover"]) if is_number(record.get("turnover")) else None,
                    }
                )
            if rows:
                return rows
        except Exception:
            continue
    return []


def fetch_json(url: str) -> dict[str, Any]:
    last_error: Exception | None = None
    try:
        return fetch_json_with_requests(url)
    except Exception as exc:  # noqa: BLE001
        last_error = exc

    for attempt in range(1, 4):
        try:
            request = urllib.request.Request(
                url,
                headers={
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
                    "Referer": "https://quote.eastmoney.com/",
                    "Accept": "application/json,text/plain,*/*",
                },
            )
            context = ssl.create_default_context()
            with urllib.request.urlopen(request, timeout=25, context=context) as response:
                return json.loads(response.read().decode("utf-8"))
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            time.sleep(attempt)

    if shutil.which("powershell.exe"):
        try:
            return fetch_json_with_powershell(url)
        except Exception as exc:  # noqa: BLE001
            raise RuntimeError(f"Failed to fetch url after retries: {last_error}; powershell: {exc}") from exc

    raise RuntimeError(f"Failed to fetch url after retries: {last_error}")


def fetch_json_with_requests(url: str) -> dict[str, Any]:
    import requests

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/125.0 Safari/537.36",
        "Referer": "https://quote.eastmoney.com/",
        "Accept": "application/json,text/plain,*/*",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
        "Connection": "close",
    }
    last_error: Exception | None = None
    for attempt in range(1, 4):
        try:
            response = requests.get(url, headers=headers, timeout=25)
            response.raise_for_status()
            return response.json()
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            time.sleep(attempt)
    raise RuntimeError(f"requests fetch failed: {last_error}")


def fetch_json_with_powershell(url: str) -> dict[str, Any]:
    command = (
        "$ErrorActionPreference='Stop'; "
        "$ProgressPreference='SilentlyContinue'; "
        f"$r=Invoke-RestMethod -Uri '{url}' -TimeoutSec 25; "
        "$r | ConvertTo-Json -Depth 8 -Compress"
    )
    result = subprocess.run(
        ["powershell.exe", "-NoProfile", "-Command", command],
        check=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=35,
    )
    return json.loads(result.stdout)


def trend_state(close: float, ma5: float | None, ma20: float | None, ma60: float | None, prev_ma20: float | None) -> str:
    if not all(is_number(v) for v in [close, ma5, ma20, ma60]):
        return "数据不足"
    ma5_f, ma20_f, ma60_f = float(ma5), float(ma20), float(ma60)
    ma20_up = is_number(prev_ma20) and ma20_f > float(prev_ma20)
    if close > ma5_f > ma20_f > ma60_f:
        return "强势"
    if close < ma60_f:
        return "弱势"
    if close > ma20_f and ma20_up:
        return "偏强"
    if abs(close / ma20_f - 1) <= 0.03:
        return "震荡"
    if close < ma20_f:
        return "偏弱"
    return "震荡"


def env_state(close: float, ma20: float | None, ma60: float | None, ret20: float | None) -> str:
    if not all(is_number(v) for v in [close, ma20, ma60]):
        return "数据不足"
    if close < float(ma60) or (is_number(ret20) and float(ret20) <= -0.05):
        return "逆风"
    if close > float(ma20) and is_number(ret20) and float(ret20) > 0:
        return "顺风"
    return "震荡"


def analyze_item(item: Item) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    rows = fetch_klines(item)
    closes = [row["close"] for row in rows]
    volumes = [row["volume"] for row in rows]
    enriched_rows: list[dict[str, Any]] = []
    for idx, row in enumerate(rows):
        close_window = closes[: idx + 1]
        volume_window = volumes[: idx + 1]
        row_ma5 = tail_avg(close_window, 5)
        row_ma20 = tail_avg(close_window, 20)
        row_ma60 = tail_avg(close_window, 60)
        row_prev_ma20 = tail_avg(close_window[:-1], 20)
        row_ret20 = row["close"] / closes[idx - 20] - 1 if idx >= 20 else None
        row_avg20vol = tail_avg(volume_window, 20)
        row_volume_ratio = row["volume"] / row_avg20vol if row_avg20vol else None
        row_trend = trend_state(row["close"], row_ma5, row_ma20, row_ma60, row_prev_ma20)
        row_signal = daily_signal(row["pct_change"], row["close"], row_ma20, row_ma60, row_volume_ratio, row_trend)
        enriched = dict(row)
        enriched.update(
            {
                "MA5": row_ma5,
                "MA20": row_ma20,
                "MA60": row_ma60,
                "20日涨跌幅": row_ret20,
                "20日平均成交量": row_avg20vol,
                "成交量倍率": row_volume_ratio,
                "趋势状态": row_trend,
                "日线标记": row_signal,
            }
        )
        enriched_rows.append(enriched)

    last = rows[-1]
    ma5 = tail_avg(closes, 5)
    ma20 = tail_avg(closes, 20)
    ma60 = tail_avg(closes, 60)
    prev_ma20 = tail_avg(closes[:-1], 20)
    avg20vol = tail_avg(volumes, 20)
    ret20 = last["close"] / closes[-21] - 1 if len(closes) > 20 else None
    volume_ratio = last["volume"] / avg20vol if avg20vol else None
    down3 = (
        len(closes) >= 4
        and closes[-1] < closes[-2]
        and closes[-2] < closes[-3]
        and closes[-3] < closes[-4]
    )
    trend = trend_state(last["close"], ma5, ma20, ma60, prev_ma20)
    env = env_state(last["close"], ma20, ma60, ret20)
    buy_return = last["close"] / item.buy_price - 1 if item.buy_price else None
    summary = {
        "类别": item.kind,
        "代码": item.code,
        "名称": item.name,
        "对应板块": item.sector,
        "日期": last["date"],
        "收盘价": last["close"],
        "今日涨跌幅": last["pct_change"],
        "MA5": ma5,
        "MA20": ma20,
        "MA60": ma60,
        "收盘-MA5": last["close"] / ma5 - 1 if ma5 else None,
        "收盘-MA20": last["close"] / ma20 - 1 if ma20 else None,
        "收盘-MA60": last["close"] / ma60 - 1 if ma60 else None,
        "20日涨跌幅": ret20,
        "今日成交量": last["volume"],
        "20日平均成交量": avg20vol,
        "成交量倍率": volume_ratio,
        "趋势状态": trend,
        "环境判断": env,
        "连续3天下跌": "是" if down3 else "否",
        "买入价": item.buy_price,
        "较买入价浮动": buy_return,
    }
    ma_series = {
        "MA5": [row["MA5"] for row in enriched_rows if is_number(row["MA5"])],
        "MA20": [row["MA20"] for row in enriched_rows if is_number(row["MA20"])],
        "MA60": [row["MA60"] for row in enriched_rows if is_number(row["MA60"])],
    }
    for name, series in ma_series.items():
        recent20 = series[-20:] if len(series) >= 20 else series
        recent60 = series[-60:] if len(series) >= 60 else series
        summary[f"{name}近20日高点"] = max(recent20) if recent20 else None
        summary[f"{name}近20日低点"] = min(recent20) if recent20 else None
        summary[f"{name}近60日高点"] = max(recent60) if recent60 else None
        summary[f"{name}近60日低点"] = min(recent60) if recent60 else None
        current = summary[name]
        summary[f"{name}距20日高点"] = current / summary[f"{name}近20日高点"] - 1 if is_number(current) and is_number(summary[f"{name}近20日高点"]) else None
        summary[f"{name}距20日低点"] = current / summary[f"{name}近20日低点"] - 1 if is_number(current) and is_number(summary[f"{name}近20日低点"]) else None
        summary[f"{name}距60日高点"] = current / summary[f"{name}近60日高点"] - 1 if is_number(current) and is_number(summary[f"{name}近60日高点"]) else None
        summary[f"{name}距60日低点"] = current / summary[f"{name}近60日低点"] - 1 if is_number(current) and is_number(summary[f"{name}近60日低点"]) else None

    for row in enriched_rows:
        row.update({"代码": item.code, "名称": item.name, "类别": item.kind})
    return summary, enriched_rows


def daily_signal(
    pct_change: float | None,
    close: float,
    ma20: float | None,
    ma60: float | None,
    volume_ratio: float | None,
    trend: str,
) -> str:
    if is_number(ma60) and close < float(ma60):
        return "红-跌破MA60"
    if is_number(pct_change) and pct_change <= -0.03 and is_number(volume_ratio) and float(volume_ratio) >= 1.5:
        return "红-放量大跌"
    if is_number(ma20) and close < float(ma20):
        return "黄-跌破MA20"
    if trend in {"强势", "偏强"} and is_number(pct_change) and pct_change > 0:
        return "绿-趋势向上"
    if trend == "震荡":
        return "黄-震荡观察"
    return "无"


def decide_action(stock: dict[str, Any], sector: dict[str, Any], market_state: str) -> tuple[str, str]:
    reasons: list[str] = []
    high = False
    medium = False

    if stock["收盘价"] < stock["MA60"]:
        high = True
        reasons.append("跌破MA60")
    if is_number(stock["今日涨跌幅"]) and stock["今日涨跌幅"] <= -0.03 and is_number(stock["成交量倍率"]) and stock["成交量倍率"] >= 1.5:
        high = True
        reasons.append("放量大跌")
    if is_number(stock["20日涨跌幅"]) and stock["20日涨跌幅"] <= -0.10:
        high = True
        reasons.append("20日跌幅超过10%")
    if is_number(stock["较买入价浮动"]) and stock["较买入价浮动"] <= -0.10:
        high = True
        reasons.append("较买入价回撤超过10%")

    if stock["收盘价"] < stock["MA20"]:
        medium = True
        reasons.append("跌破MA20")
    if stock["连续3天下跌"] == "是":
        medium = True
        reasons.append("连续3天下跌")
    if is_number(stock["今日涨跌幅"]) and stock["今日涨跌幅"] <= -0.03:
        medium = True
        reasons.append("今日跌幅超过3%")

    relative_sector = None
    if is_number(stock["20日涨跌幅"]) and is_number(sector.get("20日涨跌幅")):
        relative_sector = stock["20日涨跌幅"] - sector["20日涨跌幅"]
    if is_number(relative_sector) and relative_sector <= -0.05:
        medium = True
        reasons.append("20日明显弱于板块")

    if high:
        return "风险升高", "；".join(reasons)
    if medium:
        return "需要复查", "；".join(reasons)
    if sector["环境判断"] == "逆风" or market_state == "逆风":
        return "暂不加仓", "板块或大盘逆风，先观察持续性"
    if stock["趋势状态"] in {"强势", "偏强"} and stock["今日涨跌幅"] >= 0.04:
        return "只观察不追高", "单日涨幅较大，避免情绪化追高"
    return "正常观察", "未触发硬复查项"


def validate_data_freshness() -> None:
    required = [item for item in ITEMS if item.kind in {"stock", "index"}]
    missing = [item.name for item in required if item.code not in DATA_AUDIT]
    if missing:
        raise RuntimeError(f"缺少数据质量记录：{', '.join(missing)}")

    cached = [item.name for item in ITEMS if DATA_AUDIT[item.code]["是否缓存"] == "是"]
    if cached:
        raise RuntimeError(f"存在旧 Excel 缓存，拒绝发布：{', '.join(cached)}")

    stock_dates = {
        item.name: str(DATA_AUDIT[item.code]["最新行情日期"])
        for item in ITEMS
        if item.kind == "stock"
    }
    latest_date = max(stock_dates.values())
    stale = [f"{name}={date}" for name, date in stock_dates.items() if date != latest_date]
    if stale:
        raise RuntimeError(f"四只股票行情日期不一致，最新日期为 {latest_date}：{', '.join(stale)}")

    other_stale = [
        f"{item.name}={DATA_AUDIT[item.code]['最新行情日期']}"
        for item in ITEMS
        if str(DATA_AUDIT[item.code]["最新行情日期"]) != latest_date
    ]
    if other_stale:
        raise RuntimeError(f"指数或板块行情日期落后于股票日期 {latest_date}：{', '.join(other_stale)}")


def rounded_row(row: dict[str, Any]) -> dict[str, Any]:
    result = {}
    for key, value in row.items():
        if isinstance(value, float):
            result[key] = round(value, 6)
        else:
            result[key] = value
    return result


def write_sheet(ws, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])


def style_workbook(wb: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fills = {
        "风险升高": "F4CCCC",
        "需要复查": "FCE5CD",
        "暂不加仓": "FFF2CC",
        "只观察不追高": "D9EAD3",
        "正常观察": "D9EAD3",
        "顺风": "D9EAD3",
        "震荡": "FFF2CC",
        "逆风": "F4CCCC",
        "强势": "D9EAD3",
        "偏强": "E2F0D9",
        "偏弱": "FCE5CD",
        "弱势": "F4CCCC",
    }
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False
        headers = [cell.value for cell in ws[1]]
        header_index = {name: idx + 1 for idx, name in enumerate(headers)}
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            signal_value = None
            if "日线标记" in header_index:
                signal_value = ws.cell(row=row[0].row, column=header_index["日线标记"]).value
            row_fill = None
            if isinstance(signal_value, str):
                if signal_value.startswith("红"):
                    row_fill = PatternFill("solid", fgColor="F4CCCC")
                elif signal_value.startswith("黄"):
                    row_fill = PatternFill("solid", fgColor="FFF2CC")
                elif signal_value.startswith("绿"):
                    row_fill = PatternFill("solid", fgColor="D9EAD3")
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
                if row_fill:
                    cell.fill = row_fill
                if cell.value in fills:
                    cell.fill = PatternFill("solid", fgColor=fills[cell.value])
        for idx, header in enumerate(headers, start=1):
            letter = get_column_letter(idx)
            if any(token in str(header) for token in ["涨跌幅", "浮动", "收盘-MA", "相对", "距"]) or str(header) in {"pct_change", "amplitude", "turnover"}:
                for cell in ws[letter][1:]:
                    cell.number_format = "0.00%"
            elif str(header) in {
                "收盘价",
                "MA5",
                "MA20",
                "MA60",
                "买入价",
                "当前均线值",
                "近20日高点",
                "近20日低点",
                "近60日高点",
                "近60日低点",
                "open",
                "close",
                "high",
                "low",
            }:
                for cell in ws[letter][1:]:
                    cell.number_format = "0.00"
            elif "成交量" in str(header):
                for cell in ws[letter][1:]:
                    cell.number_format = "#,##0"
            elif str(header) == "成交量倍率":
                for cell in ws[letter][1:]:
                    cell.number_format = "0.00"
        for col_idx, cells in enumerate(ws.columns, start=1):
            width = 10
            for cell in cells:
                width = max(width, min(len(str(cell.value or "")), 36) + 2)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        for name in ["三轮复盘", "复查原因", "明日重点"]:
            if name in header_index:
                ws.column_dimensions[get_column_letter(header_index[name])].width = 48


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    summaries: list[dict[str, Any]] = []
    history_rows: list[dict[str, Any]] = []
    for item in ITEMS:
        summary, rows = analyze_item(item)
        audit = DATA_AUDIT[item.code]
        summary.update(
            {
                "数据来源": audit["数据来源"],
                "是否缓存": audit["是否缓存"],
                "数据说明": audit["说明"],
            }
        )
        summaries.append(rounded_row(summary))
        history_rows.extend(rows)

    validate_data_freshness()

    stocks = [row for row in summaries if row["类别"] == "stock"]
    indices = [row for row in summaries if row["类别"] == "index"]
    sectors = {row["名称"]: row for row in summaries if row["类别"] == "sector"}
    market_state = "顺风" if sum(row["环境判断"] == "顺风" for row in indices) >= 2 else "震荡"
    if sum(row["环境判断"] == "逆风" for row in indices) >= 2:
        market_state = "逆风"

    overview_rows: list[dict[str, Any]] = []
    ma_rows: list[dict[str, Any]] = []
    ma_range_rows: list[dict[str, Any]] = []
    for item_summary in summaries:
        for ma_name in ["MA5", "MA20", "MA60"]:
            ma_range_rows.append(
                {
                    "类别": item_summary["类别"],
                    "代码": item_summary["代码"],
                    "名称": item_summary["名称"],
                    "均线": ma_name,
                    "当前均线值": item_summary.get(ma_name),
                    "近20日高点": item_summary.get(f"{ma_name}近20日高点"),
                    "近20日低点": item_summary.get(f"{ma_name}近20日低点"),
                    "距20日高点": item_summary.get(f"{ma_name}距20日高点"),
                    "距20日低点": item_summary.get(f"{ma_name}距20日低点"),
                    "近60日高点": item_summary.get(f"{ma_name}近60日高点"),
                    "近60日低点": item_summary.get(f"{ma_name}近60日低点"),
                    "距60日高点": item_summary.get(f"{ma_name}距60日高点"),
                    "距60日低点": item_summary.get(f"{ma_name}距60日低点"),
                }
            )

    for stock in stocks:
        sector = sectors.get(stock["对应板块"], {})
        action, reason = decide_action(stock, sector, market_state)
        relative_sector = (
            stock["20日涨跌幅"] - sector.get("20日涨跌幅")
            if is_number(stock["20日涨跌幅"]) and is_number(sector.get("20日涨跌幅"))
            else None
        )
        overview_rows.append(
            {
                "日期": stock["日期"],
                "股票代码": stock["代码"],
                "股票名称": stock["名称"],
                "收盘价": stock["收盘价"],
                "今日涨跌幅": stock["今日涨跌幅"],
                "趋势状态": stock["趋势状态"],
                "对应板块": stock["对应板块"],
                "板块环境": sector.get("环境判断"),
                "大盘环境": market_state,
                "20日相对板块": relative_sector,
                "数据来源": stock["数据来源"],
                "是否缓存": stock["是否缓存"],
                "行动状态": action,
                "复查原因": reason,
                "明日重点": "看能否继续站稳MA20/MA60，并确认个股强弱是否得到板块配合。",
            }
        )
        ma_rows.append(
            {
                "股票代码": stock["代码"],
                "股票名称": stock["名称"],
                "收盘价": stock["收盘价"],
                "MA5": stock["MA5"],
                "MA20": stock["MA20"],
                "MA60": stock["MA60"],
                "收盘-MA5": stock["收盘-MA5"],
                "收盘-MA20": stock["收盘-MA20"],
                "收盘-MA60": stock["收盘-MA60"],
                "20日涨跌幅": stock["20日涨跌幅"],
                "成交量倍率": stock["成交量倍率"],
                "连续3天下跌": stock["连续3天下跌"],
                "买入价": stock["买入价"],
                "较买入价浮动": stock["较买入价浮动"],
            }
        )

    replay_rows = [
        {
            "轮次": "第一轮",
            "视角": "数据与环境",
            "三轮复盘": "先看事实：指数多数仍在趋势上方，但板块分化明显。半导体较稳，光学光电子、电池、通信设备均偏逆风。",
        },
        {
            "轮次": "第二轮",
            "视角": "相对强弱与矛盾",
            "三轮复盘": "三安、蔚蓝锂芯、中天科技都出现个股强于板块的背离，需要复查是基本面改善、消息驱动还是短线资金推动。",
        },
        {
            "轮次": "第三轮",
            "视角": "纪律与行动",
            "三轮复盘": "没有把上涨直接解释为可以买。板块逆风时以观察和复查为主；单日大涨只观察不追高；跌破MA20/MA60或回撤到复查线再升级处理。",
        },
    ]

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb.create_sheet("今日总览"), overview_rows)
    write_sheet(wb.create_sheet("MA数据对比"), ma_rows)
    write_sheet(wb.create_sheet("均线高低点"), ma_range_rows)
    write_sheet(wb.create_sheet("大盘板块"), indices + list(sectors.values()))
    write_sheet(wb.create_sheet("数据质量"), list(DATA_AUDIT.values()))
    write_sheet(wb.create_sheet("三轮复盘"), replay_rows)
    write_sheet(wb.create_sheet("原始日线"), [rounded_row(row) for row in history_rows])
    style_workbook(wb)
    wb.save(OUTPUT_FILE)
    print(OUTPUT_FILE)


if __name__ == "__main__":
    main()
