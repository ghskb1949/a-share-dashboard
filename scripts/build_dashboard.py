from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
OUTPUT_DIR = BASE_DIR / "output"
DASHBOARD_FILE = OUTPUT_DIR / "dashboard.html"


def latest_report() -> Path:
    reports = sorted(OUTPUT_DIR.glob("codex_daily_observation_*.xlsx"), reverse=True)
    if not reports:
        raise FileNotFoundError("No codex_daily_observation_*.xlsx report found in output.")
    return reports[0]


def clean_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return value


def read_sheet(wb, name: str) -> list[dict[str, Any]]:
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(item) if item is not None else "" for item in rows[0]]
    output: list[dict[str, Any]] = []
    for row in rows[1:]:
        record = {headers[idx]: clean_value(value) for idx, value in enumerate(row) if idx < len(headers)}
        if any(value is not None for value in record.values()):
            output.append(record)
    return output


def build_payload(report_path: Path) -> dict[str, Any]:
    wb = load_workbook(report_path, data_only=True, read_only=True)
    return {
        "meta": {
            "reportFile": report_path.name,
            "reportPath": str(report_path),
            "dashboardGeneratedAt": datetime.now(ZoneInfo("Asia/Shanghai")).strftime("%Y-%m-%d %H:%M:%S CST"),
        },
        "overview": read_sheet(wb, "今日总览"),
        "maCompare": read_sheet(wb, "MA数据对比"),
        "maRanges": read_sheet(wb, "均线高低点"),
        "marketSectors": read_sheet(wb, "大盘板块"),
        "review": read_sheet(wb, "三轮复盘"),
        "daily": read_sheet(wb, "原始日线"),
    }


def html_template(payload: dict[str, Any]) -> str:
    json_text = json.dumps(payload, ensure_ascii=False).replace("</", "<\\/")
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>A股四股观察驾驶舱</title>
  <style>
    :root {{
      --bg: #f6f7f9;
      --panel: #ffffff;
      --text: #17202a;
      --muted: #687385;
      --line: #d9e0ea;
      --blue: #285f9f;
      --teal: #16756f;
      --green-bg: #dcefdc;
      --green: #1f7a3a;
      --yellow-bg: #fff0bf;
      --yellow: #8a6200;
      --red-bg: #f8d7da;
      --red: #a4232e;
      --gray-bg: #eef1f5;
      --shadow: 0 1px 2px rgba(20, 30, 45, .08);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      background: var(--bg);
      color: var(--text);
      font-family: "Microsoft YaHei", "Segoe UI", Arial, sans-serif;
      font-size: 14px;
      letter-spacing: 0;
    }}
    header {{
      position: sticky;
      top: 0;
      z-index: 10;
      background: rgba(255, 255, 255, .96);
      border-bottom: 1px solid var(--line);
      backdrop-filter: blur(8px);
    }}
    .topbar {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 16px;
      max-width: 1480px;
      margin: 0 auto;
      padding: 12px 18px;
    }}
    h1 {{
      margin: 0;
      font-size: 20px;
      line-height: 1.25;
      font-weight: 700;
    }}
    .subtle {{ color: var(--muted); font-size: 12px; }}
    .layout {{
      max-width: 1480px;
      margin: 0 auto;
      padding: 16px 18px 28px;
      display: grid;
      grid-template-columns: 320px 1fr;
      gap: 16px;
    }}
    .side, .main {{ min-width: 0; }}
    .panel {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
      margin-bottom: 16px;
      overflow: hidden;
    }}
    .panel-head {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 12px;
      padding: 12px 14px;
      border-bottom: 1px solid var(--line);
      background: #fbfcfe;
    }}
    .panel h2 {{
      margin: 0;
      font-size: 15px;
      line-height: 1.3;
    }}
    .panel-body {{ padding: 14px; }}
    .kpis {{
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 10px;
      margin-bottom: 16px;
    }}
    .kpi {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 12px;
      box-shadow: var(--shadow);
    }}
    .kpi-value {{
      font-size: 24px;
      font-weight: 700;
      line-height: 1;
      margin-top: 8px;
    }}
    .stock-list {{ display: grid; gap: 10px; }}
    .stock-card {{
      width: 100%;
      text-align: left;
      border: 1px solid var(--line);
      background: #fff;
      border-radius: 8px;
      padding: 10px;
      cursor: pointer;
    }}
    .stock-card.active {{
      border-color: var(--blue);
      box-shadow: inset 3px 0 0 var(--blue);
    }}
    .stock-title {{
      display: flex;
      justify-content: space-between;
      gap: 8px;
      align-items: center;
      margin-bottom: 8px;
      font-weight: 700;
    }}
    .stock-metrics {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 6px 8px;
      color: var(--muted);
      font-size: 12px;
    }}
    .badge {{
      display: inline-flex;
      align-items: center;
      min-height: 22px;
      padding: 2px 8px;
      border-radius: 999px;
      font-size: 12px;
      font-weight: 700;
      white-space: nowrap;
    }}
    .badge.normal, .badge.up, .badge.green, .badge.wind {{ background: var(--green-bg); color: var(--green); }}
    .badge.watch, .badge.neutral, .badge.yellow {{ background: var(--yellow-bg); color: var(--yellow); }}
    .badge.risk, .badge.down, .badge.red {{ background: var(--red-bg); color: var(--red); }}
    .badge.gray {{ background: var(--gray-bg); color: var(--muted); }}
    .positive {{ color: var(--green); font-weight: 700; }}
    .negative {{ color: var(--red); font-weight: 700; }}
    .grid-2 {{
      display: grid;
      grid-template-columns: minmax(0, 1.1fr) minmax(360px, .9fr);
      gap: 16px;
    }}
    canvas {{
      width: 100%;
      height: 320px;
      display: block;
      background: #fff;
    }}
    .table-wrap {{
      overflow: auto;
      max-height: 520px;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      font-size: 13px;
      min-width: 760px;
    }}
    th, td {{
      border-bottom: 1px solid var(--line);
      padding: 8px 9px;
      text-align: right;
      vertical-align: top;
      white-space: nowrap;
    }}
    th {{
      position: sticky;
      top: 0;
      z-index: 1;
      background: #eef3f8;
      color: #223142;
      font-weight: 700;
    }}
    td:first-child, th:first-child,
    td:nth-child(2), th:nth-child(2) {{
      text-align: left;
    }}
    tr.row-red td {{ background: #fff1f1; }}
    tr.row-yellow td {{ background: #fff9e6; }}
    tr.row-green td {{ background: #eff8ef; }}
    .ma-list {{
      display: grid;
      gap: 10px;
    }}
    .ma-item {{
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 10px;
    }}
    .ma-row {{
      display: grid;
      grid-template-columns: 58px 1fr 72px;
      align-items: center;
      gap: 8px;
      margin-top: 8px;
      font-size: 12px;
      color: var(--muted);
    }}
    .bar {{
      height: 8px;
      border-radius: 999px;
      background: #e6ebf1;
      overflow: hidden;
    }}
    .bar span {{
      display: block;
      height: 100%;
      border-radius: inherit;
      background: var(--teal);
    }}
    .review-grid {{
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 12px;
    }}
    .review-card {{
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 12px;
      background: #fff;
    }}
    .review-card strong {{
      display: block;
      margin-bottom: 8px;
      color: var(--blue);
    }}
    .toolbar {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      align-items: center;
    }}
    .link-btn {{
      color: #fff;
      background: var(--blue);
      text-decoration: none;
      border-radius: 6px;
      padding: 7px 10px;
      font-size: 12px;
      font-weight: 700;
    }}
    .select {{
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 7px 8px;
      background: #fff;
      color: var(--text);
    }}
    @media (max-width: 1100px) {{
      .layout {{ grid-template-columns: 1fr; }}
      .kpis {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
      .grid-2 {{ grid-template-columns: 1fr; }}
      .review-grid {{ grid-template-columns: 1fr; }}
    }}
    @media (max-width: 640px) {{
      .topbar {{ align-items: flex-start; flex-direction: column; }}
      .kpis {{ grid-template-columns: 1fr; }}
      .layout {{ padding: 12px; }}
      table {{ font-size: 12px; }}
    }}
  </style>
</head>
<body>
  <header>
    <div class="topbar">
      <div>
        <h1>A股四只自选股观察驾驶舱</h1>
        <div class="subtle" id="metaText"></div>
      </div>
      <div class="toolbar">
        <select class="select" id="stockSelect" aria-label="选择股票"></select>
        <a class="link-btn" id="excelLink" href="#">打开 Excel</a>
      </div>
    </div>
  </header>
  <main class="layout">
    <aside class="side">
      <section class="panel">
        <div class="panel-head"><h2>自选股状态</h2><span class="subtle">点击切换</span></div>
        <div class="panel-body"><div class="stock-list" id="stockCards"></div></div>
      </section>
      <section class="panel">
        <div class="panel-head"><h2>大盘与板块</h2></div>
        <div class="panel-body table-wrap" id="envTable"></div>
      </section>
    </aside>
    <section class="main">
      <div class="kpis" id="kpis"></div>
      <div class="grid-2">
        <section class="panel">
          <div class="panel-head"><h2 id="chartTitle">趋势图</h2><span class="subtle">收盘价 / MA5 / MA20 / MA60</span></div>
          <div class="panel-body"><canvas id="trendCanvas"></canvas></div>
        </section>
        <section class="panel">
          <div class="panel-head"><h2>均线高低点</h2><span class="subtle">分开看 MA5 / MA20 / MA60</span></div>
          <div class="panel-body"><div class="ma-list" id="maRanges"></div></div>
        </section>
      </div>
      <section class="panel">
        <div class="panel-head"><h2>今日总览</h2></div>
        <div class="table-wrap" id="overviewTable"></div>
      </section>
      <section class="panel">
        <div class="panel-head"><h2>三轮复盘</h2></div>
        <div class="panel-body"><div class="review-grid" id="reviewGrid"></div></div>
      </section>
      <section class="panel">
        <div class="panel-head"><h2 id="dailyTitle">原始日线</h2><span class="subtle">红黄绿标记</span></div>
        <div class="table-wrap" id="dailyTable"></div>
      </section>
    </section>
  </main>
  <script id="payload" type="application/json">{json_text}</script>
  <script>
    const DATA = JSON.parse(document.getElementById('payload').textContent);
    const stocks = DATA.overview || [];
    let selectedCode = stocks[0]?.['股票代码'] || '';

    const pctFields = new Set(['今日涨跌幅','20日涨跌幅','20日相对板块','收盘-MA5','收盘-MA20','收盘-MA60','较买入价浮动','距20日高点','距20日低点','距60日高点','距60日低点','pct_change','amplitude','turnover']);

    function clsFor(value) {{
      const text = String(value || '');
      if (text.includes('风险') || text.includes('逆风') || text.includes('跌破') || text.startsWith('红')) return 'risk red';
      if (text.includes('复查') || text.includes('暂不') || text.includes('震荡') || text.startsWith('黄')) return 'watch yellow';
      if (text.includes('正常') || text.includes('强') || text.includes('顺风') || text.startsWith('绿')) return 'normal green';
      return 'gray';
    }}
    function fmt(value, field='') {{
      if (value === null || value === undefined || value === '') return '-';
      if (typeof value === 'number') {{
        if (pctFields.has(field)) return (value * 100).toFixed(2) + '%';
        if (Math.abs(value) >= 1000000) return Math.round(value).toLocaleString('zh-CN');
        if (Math.abs(value) >= 1000) return value.toLocaleString('zh-CN', {{maximumFractionDigits: 2}});
        return value.toFixed(2);
      }}
      return value;
    }}
    function pctClass(value) {{
      if (typeof value !== 'number') return '';
      if (value > 0) return 'positive';
      if (value < 0) return 'negative';
      return '';
    }}
    function badge(text) {{
      return `<span class="badge ${{clsFor(text)}}">${{text || '-'}}</span>`;
    }}
    function renderKpis() {{
      const risk = stocks.filter(s => String(s['行动状态']).includes('风险')).length;
      const review = stocks.filter(s => String(s['行动状态']).includes('复查')).length;
      const pause = stocks.filter(s => String(s['行动状态']).includes('暂不')).length;
      const normal = stocks.filter(s => String(s['行动状态']).includes('正常')).length;
      const items = [
        ['风险升高', risk, '需要优先处理'],
        ['需要复查', review, '检查触发原因'],
        ['暂不加仓', pause, '环境或板块不配合'],
        ['正常观察', normal, '未触发硬风险']
      ];
      document.getElementById('kpis').innerHTML = items.map(([label, value, note]) => `
        <div class="kpi"><div class="subtle">${{label}}</div><div class="kpi-value">${{value}}</div><div class="subtle">${{note}}</div></div>
      `).join('');
    }}
    function renderStockControls() {{
      const select = document.getElementById('stockSelect');
      select.innerHTML = stocks.map(s => `<option value="${{s['股票代码']}}">${{s['股票名称']}} ${{s['股票代码']}}</option>`).join('');
      select.value = selectedCode;
      select.onchange = () => {{ selectedCode = select.value; renderAll(); }};
      document.getElementById('stockCards').innerHTML = stocks.map(s => `
        <button class="stock-card ${{s['股票代码'] === selectedCode ? 'active' : ''}}" data-code="${{s['股票代码']}}">
          <div class="stock-title"><span>${{s['股票名称']}} <span class="subtle">${{s['股票代码']}}</span></span>${{badge(s['行动状态'])}}</div>
          <div class="stock-metrics">
            <span>收盘 ${{fmt(s['收盘价'])}}</span>
            <span class="${{pctClass(s['今日涨跌幅'])}}">今日 ${{fmt(s['今日涨跌幅'], '今日涨跌幅')}}</span>
            <span>趋势 ${{s['趋势状态']}}</span>
            <span>板块 ${{s['板块环境']}}</span>
          </div>
        </button>
      `).join('');
      document.querySelectorAll('.stock-card').forEach(btn => {{
        btn.addEventListener('click', () => {{ selectedCode = btn.dataset.code; renderAll(); }});
      }});
    }}
    function table(rows, fields) {{
      if (!rows.length) return '<div class="panel-body subtle">暂无数据</div>';
      const headers = fields || Object.keys(rows[0]);
      return `<table><thead><tr>${{headers.map(h => `<th>${{h}}</th>`).join('')}}</tr></thead><tbody>${{
        rows.map(row => {{
          const mark = String(row['日线标记'] || '');
          const rowClass = mark.startsWith('红') ? 'row-red' : mark.startsWith('黄') ? 'row-yellow' : mark.startsWith('绿') ? 'row-green' : '';
          return `<tr class="${{rowClass}}">${{headers.map(h => {{
            const val = row[h];
            if (h.includes('状态') || h.includes('环境') || h === '日线标记') return `<td>${{badge(val)}}</td>`;
            return `<td class="${{pctClass(val)}}">${{fmt(val, h)}}</td>`;
          }}).join('')}}</tr>`;
        }}).join('')
      }}</tbody></table>`;
    }}
    function renderOverview() {{
      document.getElementById('overviewTable').innerHTML = table(stocks, ['股票代码','股票名称','收盘价','今日涨跌幅','趋势状态','对应板块','板块环境','大盘环境','20日相对板块','行动状态','复查原因','明日重点']);
    }}
    function renderEnv() {{
      const rows = DATA.marketSectors || [];
      document.getElementById('envTable').innerHTML = table(rows, ['类别','代码','名称','收盘价','今日涨跌幅','MA20','MA60','20日涨跌幅','趋势状态','环境判断']);
    }}
    function renderReview() {{
      document.getElementById('reviewGrid').innerHTML = (DATA.review || []).map(r => `
        <div class="review-card"><strong>${{r['轮次']}}：${{r['视角']}}</strong><div>${{r['三轮复盘']}}</div></div>
      `).join('');
    }}
    function renderMaRanges() {{
      const rows = (DATA.maRanges || []).filter(r => r['代码'] === selectedCode);
      document.getElementById('maRanges').innerHTML = rows.map(r => {{
        const pos = Math.max(0, Math.min(100, ((Number(r['距60日低点']) || 0) / (((Number(r['距60日低点']) || 0) - (Number(r['距60日高点']) || -1)) || 1)) * 100));
        return `<div class="ma-item">
          <strong>${{r['均线']}}</strong>
          <div class="ma-row"><span>当前</span><div class="bar"><span style="width:${{pos.toFixed(0)}}%"></span></div><span>${{fmt(r['当前均线值'])}}</span></div>
          <div class="ma-row"><span>20日</span><span>高 ${{fmt(r['近20日高点'])}} / 低 ${{fmt(r['近20日低点'])}}</span><span>${{fmt(r['距20日高点'], '距20日高点')}}</span></div>
          <div class="ma-row"><span>60日</span><span>高 ${{fmt(r['近60日高点'])}} / 低 ${{fmt(r['近60日低点'])}}</span><span>${{fmt(r['距60日高点'], '距60日高点')}}</span></div>
        </div>`;
      }}).join('');
    }}
    function renderDaily() {{
      const rows = (DATA.daily || []).filter(r => r['代码'] === selectedCode).slice(-90).reverse();
      const stock = stocks.find(s => s['股票代码'] === selectedCode);
      document.getElementById('dailyTitle').textContent = `${{stock?.['股票名称'] || ''}} 原始日线`;
      document.getElementById('dailyTable').innerHTML = table(rows, ['date','close','pct_change','MA5','MA20','MA60','20日涨跌幅','成交量倍率','趋势状态','日线标记']);
    }}
    function drawChart() {{
      const canvas = document.getElementById('trendCanvas');
      const ctx = canvas.getContext('2d');
      const dpr = window.devicePixelRatio || 1;
      const rect = canvas.getBoundingClientRect();
      canvas.width = Math.round(rect.width * dpr);
      canvas.height = Math.round(rect.height * dpr);
      ctx.scale(dpr, dpr);
      ctx.clearRect(0, 0, rect.width, rect.height);
      const rows = (DATA.daily || []).filter(r => r['代码'] === selectedCode).slice(-120);
      const stock = stocks.find(s => s['股票代码'] === selectedCode);
      document.getElementById('chartTitle').textContent = `${{stock?.['股票名称'] || ''}} 趋势图`;
      if (!rows.length) return;
      const series = ['close','MA5','MA20','MA60'];
      const values = rows.flatMap(r => series.map(k => Number(r[k])).filter(Number.isFinite));
      const min = Math.min(...values);
      const max = Math.max(...values);
      const pad = 26;
      const x = i => pad + i * ((rect.width - pad * 2) / Math.max(1, rows.length - 1));
      const y = v => rect.height - pad - ((v - min) / Math.max(0.01, max - min)) * (rect.height - pad * 2);
      ctx.strokeStyle = '#d9e0ea';
      ctx.lineWidth = 1;
      for (let i = 0; i < 4; i++) {{
        const yy = pad + i * ((rect.height - pad * 2) / 3);
        ctx.beginPath(); ctx.moveTo(pad, yy); ctx.lineTo(rect.width - pad, yy); ctx.stroke();
      }}
      const colors = {{close:'#17202a', MA5:'#16756f', MA20:'#285f9f', MA60:'#a4232e'}};
      for (const key of series) {{
        ctx.beginPath();
        ctx.strokeStyle = colors[key];
        ctx.lineWidth = key === 'close' ? 2.4 : 1.6;
        rows.forEach((r, i) => {{
          const v = Number(r[key]);
          if (!Number.isFinite(v)) return;
          const xx = x(i), yy = y(v);
          if (i === 0) ctx.moveTo(xx, yy); else ctx.lineTo(xx, yy);
        }});
        ctx.stroke();
      }}
      ctx.fillStyle = '#687385';
      ctx.font = '12px Microsoft YaHei, Segoe UI, Arial';
      ctx.fillText(`高 ${{max.toFixed(2)}}`, pad, 16);
      ctx.fillText(`低 ${{min.toFixed(2)}}`, pad, rect.height - 8);
      [['收盘','#17202a'],['MA5','#16756f'],['MA20','#285f9f'],['MA60','#a4232e']].forEach(([label, color], i) => {{
        ctx.fillStyle = color;
        ctx.fillRect(rect.width - 190 + i * 48, 10, 16, 3);
        ctx.fillText(label, rect.width - 170 + i * 48, 15);
      }});
    }}
    function renderMeta() {{
      document.getElementById('metaText').textContent = `来源：${{DATA.meta.reportFile}} ｜ 生成：${{DATA.meta.dashboardGeneratedAt}}`;
      document.getElementById('excelLink').href = DATA.meta.reportFile;
    }}
    function renderAll() {{
      renderMeta();
      renderKpis();
      renderStockControls();
      renderOverview();
      renderEnv();
      renderReview();
      renderMaRanges();
      renderDaily();
      requestAnimationFrame(drawChart);
    }}
    window.addEventListener('resize', drawChart);
    renderAll();
  </script>
</body>
</html>"""


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    report_path = latest_report()
    payload = build_payload(report_path)
    DASHBOARD_FILE.write_text(html_template(payload), encoding="utf-8")
    print(DASHBOARD_FILE)


if __name__ == "__main__":
    main()
